import openpyxl
from openpyxl.styles import Font, Alignment,Border,Side
import win32com.client
from pywintypes import com_error
percentage=[0,0,0,0,0,0,0,0,0,0]
per_sum=0
Overall_percentage=0
c1=0
c2=0
c3=0
c4=0
c5=0

new_location=''
location=input('Enter Location of file: ')
for i in location:
    if (i=='\\'):
        i='/'
    new_location=new_location + i
file_name=input('Enter file name: ')
new_location2= new_location + '/' + file_name + '.xlsx'
wb = openpyxl.load_workbook(new_location2)
number_of_responses= wb['Form Responses 1'].max_row-1
print(number_of_responses)


#processing the excel data
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=2).value=='Unsatisfactory'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=2).value=='Average'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=2).value=='Good'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=2).value=='Very Good'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=2).value=='Excellent'):
        c5=c5+1
percentage[0] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=3).value=='Unsatisfactory'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=3).value=='Average'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=3).value=='Good'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=3).value=='Very Good'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=3).value=='Excellent'):
        c5=c5+1
percentage[1] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=4).value=='Unsatisfactory'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=4).value=='Average'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=4).value=='Good'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=4).value=='Very Good'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=4).value=='Excellent'):
        c5=c5+1
percentage[2] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=5).value=='Unsatisfactory'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=5).value=='Average'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=5).value=='Good'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=5).value=='Very Good'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=5).value=='Excellent'):
        c5=c5+1
percentage[3] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=6).value=='Unsatisfactory'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=6).value=='Average'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=6).value=='Good'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=6).value=='Very Good'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=6).value=='Excellent'):
        c5=c5+1
percentage[4] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=7).value=='Unsatisfactory'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=7).value=='Average'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=7).value=='Good'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=7).value=='Very Good'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=7).value=='Excellent'):
        c5=c5+1
percentage[5] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=8).value=='Negative'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=8).value=='Indifferent'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=8).value=='Dictating'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=8).value=='Friendly'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=8).value=='Positive and Helpful'):
        c5=c5+1
percentage[6] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=9).value=='Discourages'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=9).value=='Never'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=9).value=='Rarely'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=9).value=='Quite Often'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=9).value=='Always'):
        c5=c5+1
percentage[7] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=10).value=='Discourages'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=10).value=='Never'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=10).value=='Rarely'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=10).value=='Quite often'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=10).value=='Always'):
        c5=c5+1
percentage[8] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

c1=0
c2=0
c3=0
c4=0
c5=0
for rowNum in range(2, wb['Form Responses 1'].max_row+1):  
    if(wb['Form Responses 1'].cell(row=rowNum, column=11).value=='Discourages'):
        c1=c1+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=11).value=='Never'):
        c2=c2+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=11).value=='Rarely'):
        c3=c3+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=11).value=='Quite Often'):
        c4=c4+1
    if(wb['Form Responses 1'].cell(row=rowNum, column=11).value=='Always'):
        c5=c5+1
percentage[9] = round((((c1*1)+(c2*2)+(c3*3)+(c4*4)+(c5*5))/(number_of_responses*5))*100,2)

for k in percentage:
    per_sum = per_sum+k

Overall_percentage=round(per_sum/10, 2)
#End of data processing



#Creating the output excel file(feedback report) according to template.
wb2=openpyxl.Workbook()
sheet=wb2.active
sheet.column_dimensions['B'].width = 90
sheet.column_dimensions['A'].width = 7
sheet.column_dimensions['C'].width = 22
sheet.cell(row = 1, column = 1).value = 'UNIVERSITY'
sheet.cell(row = 1, column = 1).font = Font(size = 12, bold=True)
sheet.cell(row = 1, column = 1).alignment = Alignment(horizontal="center")
sheet.merge_cells('A1:C1')
sheet.cell(row = 2, column = 1).value = 'XYZ COLLEGE'
sheet.cell(row = 2, column = 1).font = Font(size = 12, bold=True)
sheet.cell(row = 2, column = 1).alignment = Alignment(horizontal="center")
sheet.merge_cells('A2:C2')
sheet.cell(row = 3, column = 1).value = 'Feedback Report   Academic Year 2020-21   Semester ODD'
sheet.cell(row = 3, column = 1).font = Font(size = 11)
sheet.cell(row = 3, column = 1).alignment = Alignment(horizontal="center")
sheet.merge_cells('A3:C3')
Name_faculty = input('Enter Name of faculty: ')
Name_Subject = input('Enter Name of Subject: ')
Name_Class = input('Enter Name of Class: ') 
Name_Semester = input('Enter Name of Semester: ')
sheet.cell(row = 4, column = 2).value = 'Name of the Faculty : '+ Name_faculty
sheet.cell(row = 4, column = 2).font = Font(size = 11)
sheet.cell(row = 5, column = 2).value = 'Name of the Subject : '+Name_Subject
sheet.cell(row = 5, column = 2).font = Font(size = 11)
sheet.cell(row = 6, column = 2).value = 'Class : ' + Name_Class
sheet.cell(row = 6, column = 2).font = Font(size = 11)
sheet.cell(row = 7, column = 2).value = 'Semester : ' + Name_Semester
sheet.cell(row = 7, column = 2).font = Font(size = 11)
sheet.cell(row = 8, column = 2).value = 'Number of Responses : '+ str(number_of_responses)
sheet.cell(row = 8, column = 2).font = Font(size = 10, bold=True)
sheet.cell(row = 9, column = 1).value = 'Sr. No'
sheet.cell(row = 9, column = 1).font = Font(size = 10, bold=True)
sheet.cell(row = 9, column = 2).value = 'Questions'
sheet.cell(row = 9, column = 2).font = Font(size = 10, bold=True)
sheet.cell(row = 9, column = 3).value = 'Aggregate Percentage of Each Question'
sheet.cell(row = 9, column = 3).font = Font(size = 10, bold=True)
sheet.cell(row = 9, column = 3).alignment = Alignment(wrap_text=True)
sheet.cell(row = 10, column = 1).value = 1
sheet.cell(row = 10, column = 1).font = Font(size = 10)
sheet.cell(row = 10, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 11, column = 1).value = 2
sheet.cell(row = 11, column = 1).font = Font(size = 10)
sheet.cell(row = 11, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 12, column = 1).value = 3
sheet.cell(row = 12, column = 1).font = Font(size = 10)
sheet.cell(row = 12, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 13, column = 1).value = 4
sheet.cell(row = 13, column = 1).font = Font(size = 10)
sheet.cell(row = 13, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 14, column = 1).value = 5
sheet.cell(row = 14, column = 1).font = Font(size = 10)
sheet.cell(row = 14, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 15, column = 1).value = 6
sheet.cell(row = 15, column = 1).font = Font(size = 10)
sheet.cell(row = 15, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 16, column = 1).value = 7
sheet.cell(row = 16, column = 1).font = Font(size = 10)
sheet.cell(row = 16, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 17, column = 1).value = 8
sheet.cell(row = 17, column = 1).font = Font(size = 10)
sheet.cell(row = 17, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 18, column = 1).value = 9
sheet.cell(row = 18, column = 1).font = Font(size = 10)
sheet.cell(row = 18, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 19, column = 1).value = 10
sheet.cell(row = 19, column = 1).font = Font(size = 10)
sheet.cell(row = 19, column = 1).alignment = Alignment(horizontal="center")
sheet.cell(row = 10, column = 2).value = 'Knowledge related to the subject (theory / Practical)'
sheet.cell(row = 10, column = 2).font = Font(size = 10)
sheet.cell(row = 11, column = 2).value = 'Delivery of lectures and pace at which material was covered?'
sheet.cell(row = 11, column = 2).font = Font(size = 10)
sheet.cell(row = 12, column = 2).value = 'Interaction with students in online class / class rooms / laboratory / staff room'
sheet.cell(row = 12, column = 2).font = Font(size = 10)
sheet.cell(row = 13, column = 2).value = 'Regularity and punctuality in conducting class / laboratory'
sheet.cell(row = 13, column = 2).font = Font(size = 10)
sheet.cell(row = 14, column = 2).value = 'Clarifying student’s doubts and availability of the instructor via email or any other online platform'
sheet.cell(row = 14, column = 2).font = Font(size = 10)
sheet.cell(row = 15, column = 2).value = 'Rate the conduction of class tests/assignments'
sheet.cell(row = 15, column = 2).font = Font(size = 10)
sheet.cell(row = 16, column = 2).value = 'Attitude towards students in online class / laboratory'
sheet.cell(row = 16, column = 2).font = Font(size = 10)
sheet.cell(row = 17, column = 2).value = 'Grades / marks for exams were informed in a timely fashion'
sheet.cell(row = 17, column = 2).font = Font(size = 10)
sheet.cell(row = 18, column = 2).value = 'Motivation for GATE / IES /IAS /PSCs / GRE / GMAT/ CAT /TOFEL/ Conferences/ Seminars / Trainings / Placements/ Projects/ Uses ICT for Teaching etc'
sheet.cell(row = 18, column = 2).font = Font(size = 10)
sheet.cell(row = 18, column = 2).alignment = Alignment(wrap_text=True)
sheet.cell(row = 19, column = 2).value = 'Rate the tools / LMS used for online teaching / laboratory'
sheet.cell(row = 19, column = 2).font = Font(size = 10)
sheet.cell(row = 10, column = 3).value = percentage[0]
sheet.cell(row = 10, column = 3).font = Font(size = 10)
sheet.cell(row = 10, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 11, column = 3).value = percentage[1]
sheet.cell(row = 11, column = 3).font = Font(size = 10)
sheet.cell(row = 11, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 12, column = 3).value = percentage[2]
sheet.cell(row = 12, column = 3).font = Font(size = 10)
sheet.cell(row = 12, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 13, column = 3).value = percentage[3]
sheet.cell(row = 13, column = 3).font = Font(size = 10)
sheet.cell(row = 13, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 14, column = 3).value = percentage[4]
sheet.cell(row = 14, column = 3).font = Font(size = 10)
sheet.cell(row = 14, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 15, column = 3).value = percentage[5]
sheet.cell(row = 15, column = 3).font = Font(size = 10)
sheet.cell(row = 15, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 16, column = 3).value = percentage[6]
sheet.cell(row = 16, column = 3).font = Font(size = 10)
sheet.cell(row = 16, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 17, column = 3).value = percentage[7]
sheet.cell(row = 17, column = 3).font = Font(size = 10)
sheet.cell(row = 17, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 18, column = 3).value = percentage[8]
sheet.cell(row = 18, column = 3).font = Font(size = 10)
sheet.cell(row = 18, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 19, column = 3).value = percentage[9]
sheet.cell(row = 19, column = 3).font = Font(size = 10)
sheet.cell(row = 19, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 21, column = 2).value = 'Overall Percentage'
sheet.cell(row = 21, column = 2).font = Font(size = 10, bold=True)
sheet.cell(row = 21, column = 3).value = str(Overall_percentage) + '%'
sheet.cell(row = 21, column = 3).font = Font(size = 10, bold=True)
sheet.cell(row = 21, column = 3).alignment = Alignment(horizontal="center")
sheet.cell(row = 23, column = 2).value = 'Teaching Efficiency Percentage'
sheet.cell(row = 23, column = 2).font = Font(size = 10, bold=True)
sheet.cell(row = 24, column = 2).value = 'Above 85% : Excellent'
sheet.cell(row = 24, column = 2).font = Font(size = 10)
sheet.cell(row = 25, column = 2).value = '85-70% : Very Good'
sheet.cell(row = 25, column = 2).font = Font(size = 10)
sheet.cell(row = 26, column = 2).value = '70-50% : Good'
sheet.cell(row = 26, column = 2).font = Font(size = 10)
sheet.cell(row = 27, column = 2).value = '50-30% : Average'
sheet.cell(row = 27, column = 2).font = Font(size = 10)
sheet.cell(row = 28, column = 2).value = 'Less than 30% : Below Average'
sheet.cell(row = 28, column = 2).font = Font(size = 10)
sheet.merge_cells('C28:D28')
sheet.cell(row = 28, column = 3).value = 'Principal\'s Signature and Stamp'
sheet.cell(row = 28, column = 3).font = Font(size = 10, bold=True)


def set_border(ws, cell_range):   #this function applies borders to all cells in the given range
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
set_border(sheet, 'A9:C19')
output_file_location = new_location + '/' + file_name + 'Feedback' + '.xlsx'
wb2.save(output_file_location)



#Generating pdf from the newly created excel report
from win32com import client
#file_name=input('Enter the file name:')
final_location= new_location + '/' + file_name+'Feedback'+'.xlsx'

PATH_TO_PDF = new_location + '/' + file_name+'Feedback'+'.pdf'
# Open Microsoft Excel
excel = client.Dispatch("Excel.Application")

# Read Excel File
sheets = excel.Workbooks.Open(final_location)
ws = sheets.Worksheets('Sheet')
ws.PageSetup.Orientation = 2
work_sheets = sheets.Worksheets[0]

# Convert into PDF File
work_sheets.ExportAsFixedFormat(0, PATH_TO_PDF)
sheets.Close()
excel.Quit()
